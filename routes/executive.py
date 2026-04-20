"""Executive intelligence dashboard + exports"""
import io
import json
import logging
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from bson import ObjectId
from datetime import datetime, timezone
from models import UserRole
from database import db
from auth_utils import require_roles
from constants import CURRENT_CBL_RATES, ARABIC_MONTHS
from helpers import ar
from pdf_generator import ensure_fonts
import openpyxl
from openpyxl.styles import PatternFill, Font as XLFont, Alignment as XLAlignment
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.colors import HexColor

router = APIRouter(prefix="/executive", tags=["executive"])
logger = logging.getLogger(__name__)


@router.get("/dashboard")
async def executive_dashboard(current_user=Depends(require_roles(
    UserRole.ADMIN, UserRole.ACID_REVIEWER, UserRole.ACID_RISK_OFFICER,
    UserRole.CUSTOMS_VALUER, UserRole.INSPECTOR
))):
    now = datetime.now(timezone.utc)
    total_requests = await db.acid_requests.count_documents({})
    approved = await db.acid_requests.count_documents({"status": "approved"})
    pending = await db.acid_requests.count_documents({"status": {"$in": ["submitted", "under_review"]}})
    rejected = await db.acid_requests.count_documents({"status": "rejected"})
    high_risk = await db.acid_requests.count_documents({"risk_level": "high"})
    sad_agg = await db.sad_forms.aggregate([{"$group": {
        "_id": None, "total_revenue_lyd": {"$sum": "$total_lyd"},
        "total_duty_lyd": {"$sum": "$customs_duty_lyd"},
        "total_vat_lyd": {"$sum": "$vat_lyd"},
        "total_value_usd": {"$sum": "$value_usd"}, "count": {"$sum": 1}
    }}]).to_list(1)
    rev = sad_agg[0] if sad_agg else {}
    six_ago = datetime(now.year, max(1, now.month - 5), 1, tzinfo=timezone.utc)
    monthly = await db.acid_requests.aggregate([
        {"$match": {"created_at": {"$gte": six_ago}}},
        {"$group": {"_id": {"year": {"$year": "$created_at"}, "month": {"$month": "$created_at"}},
                    "count": {"$sum": 1},
                    "high_risk": {"$sum": {"$cond": [{"$eq": ["$risk_level", "high"]}, 1, 0]}},
                    "total_value": {"$sum": "$value_usd"}}},
        {"$sort": {"_id.year": 1, "_id.month": 1}}
    ]).to_list(12)
    ports = await db.acid_requests.aggregate([
        {"$group": {"_id": "$port_of_entry", "count": {"$sum": 1},
                    "approved": {"$sum": {"$cond": [{"$eq": ["$status", "approved"]}, 1, 0]}},
                    "high_risk": {"$sum": {"$cond": [{"$eq": ["$risk_level", "high"]}, 1, 0]}},
                    "total_value": {"$sum": "$value_usd"}}},
        {"$sort": {"count": -1}}, {"$limit": 8}
    ]).to_list(8)
    risk_dist = await db.acid_requests.aggregate([{"$group": {"_id": "$risk_level", "count": {"$sum": 1}}}]).to_list(5)
    status_dist = await db.acid_requests.aggregate([{"$group": {"_id": "$status", "count": {"$sum": 1}}}]).to_list(10)
    transport_dist = await db.acid_requests.aggregate([{"$group": {"_id": "$transport_mode", "count": {"$sum": 1}}}]).to_list(5)
    countries = await db.acid_requests.aggregate([
        {"$group": {"_id": "$supplier_country", "count": {"$sum": 1}, "total_value": {"$sum": "$value_usd"}}},
        {"$sort": {"count": -1}}, {"$limit": 8}
    ]).to_list(8)
    # Admin intelligence: guarantees + violations
    guarantee_stats = (await db.guarantees.aggregate([
        {"$match": {"status": "active"}},
        {"$group": {"_id": None, "total_amount": {"$sum": "$amount_lyd"}, "count": {"$sum": 1}}}
    ]).to_list(1) or [{"total_amount": 0, "count": 0}])[0]
    violation_fines = (await db.violations.aggregate([
        {"$match": {"status": "fined"}},
        {"$group": {"_id": None, "total_fines": {"$sum": "$fine_amount_lyd"}, "count": {"$sum": 1}}}
    ]).to_list(1) or [{"total_fines": 0, "count": 0}])[0]

    # Phase G — Platform Revenue (NAFIDHA internal fees)
    fee_agg = await db.platform_fees.aggregate([
        {"$group": {
            "_id": "$fee_type",
            "total_lyd": {"$sum": "$amount_lyd"},
            "count": {"$sum": 1},
            "paid_count": {"$sum": {"$cond": [{"$eq": ["$status", "paid"]}, 1, 0]}},
            "paid_lyd": {"$sum": {"$cond": [{"$eq": ["$status", "paid"]}, "$amount_lyd", 0]}},
        }}
    ]).to_list(20)
    fee_by_type = {f["_id"]: f for f in fee_agg}
    platform_subscription_lyd = fee_by_type.get("annual_subscription", {}).get("paid_lyd", 0) or 0
    platform_acid_fees_lyd     = fee_by_type.get("acid_fee", {}).get("paid_lyd", 0) or 0
    platform_amendment_lyd     = fee_by_type.get("amendment_fee", {}).get("paid_lyd", 0) or 0
    total_platform_revenue_lyd = round(platform_subscription_lyd + platform_acid_fees_lyd + platform_amendment_lyd, 2)
    pending_fees_count = await db.platform_fees.count_documents({"status": "pending"})
    early_bird_count   = await db.platform_fees.count_documents({"fee_type": "annual_subscription", "early_bird_discount": True})

    # Phase G — Suspended accounts count
    suspended_count = await db.users.count_documents({"is_active": False, "suspended_reason": "license_expired"})
    total_active_entities = await db.users.count_documents({
        "is_active": True, "role": {"$in": ["importer", "customs_broker", "carrier_agent"]}
    })
    total_revenue = round(rev.get("total_revenue_lyd", 0), 2)
    total_trade_usd = round(rev.get("total_value_usd", 0), 2)
    potential_revenue = round(total_trade_usd * 0.20 * CURRENT_CBL_RATES["USD"], 2)
    leakage = round(max(0, potential_revenue - total_revenue), 2)
    # Phase K — Green Channel Intelligence Metrics
    gc_total    = await db.acid_requests.count_documents({"is_green_channel": True})
    gc_pending  = await db.acid_requests.count_documents({"is_green_channel": True,  "status": {"$in": ["submitted", "under_review"]}})
    gc_approved = await db.acid_requests.count_documents({"is_green_channel": True,  "status": "approved"})

    # متوسط وقت التخليص — القناة الخضراء vs. العادية
    async def _avg_clearance_hours(is_green: bool) -> float:
        docs = await db.acid_requests.find(
            {"is_green_channel": is_green, "clearance_started_at": {"$ne": None}, "clearance_completed_at": {"$ne": None}},
            {"clearance_started_at": 1, "clearance_completed_at": 1}
        ).to_list(500)
        if not docs:
            return 0.0
        total_hours = 0.0
        for d in docs:
            try:
                start = datetime.fromisoformat(d["clearance_started_at"].replace("Z", "+00:00"))
                end   = datetime.fromisoformat(d["clearance_completed_at"].replace("Z", "+00:00"))
                total_hours += (end - start).total_seconds() / 3600
            except Exception:
                pass
        return round(total_hours / len(docs), 1)

    avg_hours_green   = await _avg_clearance_hours(True)
    avg_hours_regular = await _avg_clearance_hours(False)

    # Green Channel ports map data
    gc_ports = await db.acid_requests.aggregate([
        {"$match": {"is_green_channel": True, "status": {"$in": ["submitted", "under_review"]}}},
        {"$group": {"_id": "$port_of_entry", "count": {"$sum": 1}}}
    ]).to_list(20)

    return {
        "summary": {
            "total_requests": total_requests, "approved": approved,
            "pending": pending, "rejected": rejected, "high_risk": high_risk,
            "approval_rate_pct": round(approved / max(total_requests, 1) * 100, 1),
            "revenue_collected_lyd": total_revenue,
            "customs_duty_lyd": round(rev.get("total_duty_lyd", 0), 2),
            "vat_collected_lyd": round(rev.get("total_vat_lyd", 0), 2),
            "total_trade_value_usd": total_trade_usd,
            "sad_forms_count": rev.get("count", 0),
            "revenue_leakage_estimate_lyd": leakage,
            "cbl_usd_rate": CURRENT_CBL_RATES["USD"],
            # Admin intelligence additions
            "active_guarantees_count": guarantee_stats.get("count", 0),
            "active_guarantees_total_lyd": round(guarantee_stats.get("total_amount", 0), 2),
            "violation_fines_collected_lyd": round(violation_fines.get("total_fines", 0), 2),
            "violation_fines_count": violation_fines.get("count", 0),
            # Phase G — NAFIDHA Platform Revenue
            "platform_subscription_lyd": round(platform_subscription_lyd, 2),
            "platform_acid_fees_lyd": round(platform_acid_fees_lyd, 2),
            "platform_amendment_lyd": round(platform_amendment_lyd, 2),
            "total_platform_revenue_lyd": total_platform_revenue_lyd,
            "pending_platform_fees_count": pending_fees_count,
            "early_bird_subscriptions": early_bird_count,
            # Phase G — Compliance
            "suspended_accounts_count": suspended_count,
            "active_entities_count": total_active_entities,
        },
        "monthly_trend": [
            {"month": ARABIC_MONTHS.get(m["_id"]["month"], ""), "month_num": m["_id"]["month"],
             "year": m["_id"]["year"], "requests": m["count"], "high_risk": m["high_risk"],
             "total_value_usd": round(m.get("total_value", 0), 2)}
            for m in monthly
        ],
        "port_performance": [
            {"port": p["_id"] or "غير محدد", "total": p["count"], "approved": p["approved"],
             "high_risk": p["high_risk"], "total_value_usd": round(p.get("total_value", 0), 2),
             "efficiency_pct": round(p["approved"] / max(p["count"], 1) * 100, 1)}
            for p in ports if p["_id"]
        ],
        "risk_distribution": [{"level": r["_id"], "count": r["count"]} for r in risk_dist if r["_id"]],
        "status_distribution": [{"status": s["_id"], "count": s["count"]} for s in status_dist if s["_id"]],
        "transport_distribution": [{"mode": t["_id"], "count": t["count"]} for t in transport_dist if t["_id"]],
        "top_countries": [
            {"country": c["_id"] or "غير محدد", "requests": c["count"],
             "total_value_usd": round(c.get("total_value", 0), 2)}
            for c in countries if c["_id"]
        ],
        "generated_at": datetime.now(timezone.utc).isoformat(),
        # Phase K — Green Channel metrics
        "green_channel": {
            "total": gc_total,
            "pending": gc_pending,
            "approved": gc_approved,
            "avg_clearance_hours_green": avg_hours_green,
            "avg_clearance_hours_regular": avg_hours_regular,
            "active_ports": [{"port": p["_id"], "count": p["count"]} for p in gc_ports if p["_id"]],
        }
    }
async def export_audit_excel(current_user=Depends(require_roles(UserRole.ADMIN))):
    logs = await db.audit_logs.find({}).sort("timestamp", -1).to_list(1000)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Logs"
    NAVY_FILL = PatternFill("solid", fgColor="1e3a5f")
    headers = ["التوقيت", "الإجراء", "المستخدم", "نوع المورد", "المعرّف", "عنوان IP", "التفاصيل"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = XLFont(bold=True, color="FFFFFF", size=11)
        cell.fill = NAVY_FILL
        cell.alignment = XLAlignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 20
    for row_idx, log in enumerate(logs, 2):
        ts = log.get("timestamp")
        ts_str = ts.isoformat() if isinstance(ts, datetime) else str(ts or "")
        details_str = json.dumps(log.get("details", {}), ensure_ascii=False)[:100]
        row_data = [ts_str, log.get("action",""), log.get("user_name",""), log.get("resource_type",""),
                    log.get("resource_id","")[-8:] if log.get("resource_id") else "",
                    log.get("ip_address",""), details_str]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = XLAlignment(horizontal="right")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EAF0FA")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"audit_logs_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.get("/export/dashboard-pdf")
async def export_dashboard_pdf(current_user=Depends(require_roles(
    UserRole.ADMIN, UserRole.ACID_REVIEWER, UserRole.CUSTOMS_VALUER
))):
    ensure_fonts()
    total_requests = await db.acid_requests.count_documents({})
    approved = await db.acid_requests.count_documents({"status": "approved"})
    high_risk = await db.acid_requests.count_documents({"risk_level": "high"})
    pending = await db.acid_requests.count_documents({"status": {"$in": ["submitted", "under_review"]}})
    sad_agg = await db.sad_forms.aggregate([{"$group": {"_id": None,
        "total_lyd": {"$sum": "$total_lyd"}, "duty_lyd": {"$sum": "$customs_duty_lyd"},
        "vat_lyd": {"$sum": "$vat_lyd"}}}]).to_list(1)
    rev = sad_agg[0] if sad_agg else {}
    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    NAVY = HexColor("#1e3a5f"); GOLD = HexColor("#d4a017")
    c.setFillColor(NAVY); c.rect(0, H-55*mm, W, 55*mm, fill=True, stroke=False)
    c.setFillColor(GOLD); c.rect(0, H-57*mm, W, 2*mm, fill=True, stroke=False)
    c.setFillColor(colors.white); c.setFont('AmiriBold', 16)
    c.drawCentredString(W/2, H-18*mm, ar("لوحة الذكاء التنفيذي — نافذة الجمارك الليبية"))
    c.setFont('Helvetica-Bold', 10)
    c.drawCentredString(W/2, H-28*mm, "Executive Intelligence Dashboard — Libya Customs NAFIDHA")
    c.setFont('Helvetica', 9); c.setFillColor(HexColor("#aab"))
    c.drawCentredString(W/2, H-38*mm, f"Generated: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} UTC")
    y = H - 70*mm
    kpis = [
        ("إجمالي الطلبات", str(total_requests), "Total Requests"),
        ("الطلبات المعتمدة", str(approved), "Approved Requests"),
        ("قيد المراجعة", str(pending), "Pending Review"),
        ("عالية المخاطر", str(high_risk), "High Risk"),
        ("الإيرادات المحصلة", f"{rev.get('total_lyd',0):,.0f} د.ل", "Revenue Collected (LYD)"),
        ("الرسوم الجمركية", f"{rev.get('duty_lyd',0):,.0f} د.ل", "Customs Duty (LYD)"),
        ("ضريبة القيمة المضافة", f"{rev.get('vat_lyd',0):,.0f} د.ل", "VAT Collected (LYD)"),
        ("سعر الصرف CBL", "4.87 د.ل/دولار", "CBL Exchange Rate"),
    ]
    col_w = (W - 20*mm) / 2
    for i, (la, va, en) in enumerate(kpis):
        col = i % 2; row = i // 2
        x = 10*mm + col * col_w; cy = y - row * 22*mm
        c.setFillColor(HexColor("#EAF0FA") if (i % 4 < 2) else HexColor("#F0F4FF"))
        c.roundRect(x+1*mm, cy-15*mm, col_w-3*mm, 18*mm, 3, fill=True, stroke=False)
        c.setFillColor(NAVY); c.setFont('AmiriBold', 9)
        c.drawRightString(x+col_w-4*mm, cy, ar(la))
        c.setFont('AmiriBold', 14); c.setFillColor(GOLD)
        c.drawRightString(x+col_w-4*mm, cy-9*mm, ar(va))
        c.setFillColor(HexColor("#888")); c.setFont('Helvetica', 7)
        c.drawString(x+3*mm, cy-1*mm, en)
    c.setFillColor(NAVY); c.rect(0, 0, W, 12*mm, fill=True, stroke=False)
    c.setFillColor(colors.white); c.setFont('Amiri', 7)
    c.drawCentredString(W/2, 4*mm, ar("سري — للاستخدام الداخلي فقط | CONFIDENTIAL — Internal Use Only"))
    c.save(); buf.seek(0)
    fname = f"executive_dashboard_{datetime.now(timezone.utc).strftime('%Y%m%d')}.pdf"
    return StreamingResponse(io.BytesIO(buf.getvalue()), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


# Legacy export routes (no prefix) for frontend compatibility
export_router = APIRouter(tags=["exports"])

@export_router.get("/export/audit-excel")
async def export_audit_excel_legacy(current_user=Depends(require_roles(UserRole.ADMIN))):
    return await export_audit_excel(current_user)

@export_router.get("/export/dashboard-pdf")
async def export_dashboard_pdf_legacy(current_user=Depends(require_roles(
    UserRole.ADMIN, UserRole.ACID_REVIEWER, UserRole.CUSTOMS_VALUER
))):
    return await export_dashboard_pdf(current_user)
